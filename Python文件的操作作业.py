# """
# 期末考试结束，老师将学生们的成绩保存到名字为“班级信息表”excel表中，
# 表格的第一列存储班级序号，例如有2001、2002、2003三个班，第二列存
# 储姓名，第三列存储成绩，现在需要将每个班达到90分及以上的学生找到，
# 并将相关学生的班级和姓名保存在新的excel文件中。如何使用python程序
# 解决这一问题？
# """
# import openpyxl  # 导入openpyxl库
#
# excel = openpyxl.load_workbook('班级信息表.xlsx')  # 载入excel工作表
# sheet = excel['成绩表']  # 通过索引读取excel工作表中的工作簿
# output = openpyxl.Workbook("高分学生名单.xlsx")
# output_sheet = output.active
# rows = sheet.max_row  # 读取工作簿行值
# cols = sheet.max_column  # 读取工作簿列值
#
# #设置字典result，字典的key保存班级字符串，value是一个存储成绩的列表
# result=dict()
#
# #通过变量i，从excel表格的第二行开始遍历
# for i in range(2, rows + 1):
#     #获取表格第一列班级名称存储至classroom
#     classroom=sheet.cell(i,1).value
#     #获取表格第二列学生姓名存储至name
#     name=sheet.cell(i,2).value
#     #获取表格第三列学生成绩存储至score
#     score=sheet.cell(i,3).value
#
#     #判断学生成绩是否大于等于90分
#     if score >= 90:#如果score大于等于90
#         if score not in result:
#             result[score]=list()
#         result[score].extend(classroom,name)
#         output_sheet.append()
#



from openpyxl import load_workbook
from openpyxl import Workbook

# 加载班级信息表
wb = load_workbook(filename='班级信息表.xlsx')
sheet = wb.active

# 创建一个新的工作簿和工作表
new_wb = Workbook()
new_sheet = new_wb.active

# 设置新表头
# new_sheet.append(["班级序号", "姓名"])

# 遍历原表格，筛选出符合条件的学生，并写入新表格
for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始遍历
    class_number, name, score = row
    if score >= 90:
        new_sheet.append([class_number, name])

# 保存新的Excel文件
new_wb.save("高分学生名单.xlsx")
wb.close()
new_wb.close()