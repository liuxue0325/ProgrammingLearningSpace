"""
期末考试结束，老师将学生们的成绩保存到名字为“班级信息表”excel表中，
表格的第一列存储班级序号，例如有2001、2002、2003三个班，第二列存
储姓名，第三列存储成绩，现在需要将每个班达到90分及以上的学生找到，
并将相关学生的班级和姓名保存在新的文本文件中。如何使用python程序
解决这一问题？
"""

# 问题解决思路：程序读取名为”班级信息表.xlsx“文件-->进行相关的统计--->将结果输出到另一个TXT文件

"""
Python的文件操作
    文件操作函数：
    fileObject=open(file_name,mode)
        file_name:待处理的文件名称
        mode（打开模式）：
            r：已知读方式打开文件
            w：打开一个文件用于写入。如果文件已存在，原内容会被删除。如果该文件不存在，创建新文件。
        fileObject(文件操作对象): 包含三种调用方式 
            方法名         | 调用方式                     | 说明
            write()方法    | fileObject.write(string)   | 将传入的字符串写入到对象对应的文件中
            read()方法     |  fileObject.read([count])  | 读取对象文件中count个字符，默认为全部读取
            close()方法    |  fileObject.close()        | 关闭文件释放相关资源

********************************
# 代码示例
file1=open('file1.txt','r')#以读模式打开文件file1
file2=open('file2.txt','w')#以邪魔时打开文件file2

words=file1.read()#将file1中的数据读取存储至words

for w in words:#遍历words中的每个字符
    #如果w为大小写字符
    if w>='a' and w<='z' or w>='A' and w<='Z':
        file2.write(w)#通过write函数写入file2
        
file1.close()#关闭file1
file2.close()#关闭file2
********************************
"""

"""
第三方库的安装方法
   处理execl文件需要安装openpyxl库
   进入到pip目录下打开命令行
   执行 pip install openpyxl 命令


Python基于openpyxl操作excel文件

******************************
   # 代码示例
   import openpyxl #导入openpyxl库
   excel=openpyxl.load_workbook('班级信息表.xlsx') #载入excel工作表
   sheet=excel['成绩表'] #通过索引读取excel工作表中的工作簿
   rows=sheet.max_row #读取工作簿行值
   cols=sheet.max_column #读取工作簿列值
   for i in range(1,rows+1):
        for j in range(1,cols+1):
            item=sheet.cell(i,j).value #通过cell()读出表格中的值
            print("%s\t"%(item),end='')
        print("")
   
"""

import openpyxl  # 导入openpyxl库

excel = openpyxl.load_workbook('班级信息表.xlsx')  # 载入excel工作表
sheet = excel['成绩表']  # 通过索引读取excel工作表中的工作簿
rows = sheet.max_row  # 读取工作簿行值
cols = sheet.max_column  # 读取工作簿列值

#设置字典result，字典的key保存班级字符串，value是一个存储成绩的列表
result=dict()

#通过变量i，从excel表格的第二行开始遍历
for i in range(2, rows + 1):
    #获取表格第一列班级名称存储至classroom
    classroom=sheet.cell(i,1).value
    #获取表格第二列学生姓名存储至name
    name=sheet.cell(i,2).value
    #获取表格第三列学生成绩存储至score
    score=sheet.cell(i,3).value

    #判断学生成绩是否大于等于90分
    if score >= 90:#如果score大于等于90
        if classroom not in result:#当第一次添加该成绩对应的班级时
            result[classroom]=list()#将该班级对应一个空的列表
        result[classroom].append(name)#将该成绩对应的学生名字添加到班级对应的列表中
    output=open('score.txt','w')#打开存储结果文件score.txt

    for classroom in result:#遍历结果字典result
        for name in result[classroom]:#遍历班级列表中的姓名
            output.write("%s\t%s\n"%(classroom,name))


#最后关闭输入文件excel和输出文件output
excel.close()
output.close()
